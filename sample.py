import json
from bs4 import BeautifulSoup
import requests
from openpyxl.workbook.workbook import Workbook
import datetime
import time
import openpyxl
from openpyxl.chart.reference import Reference
from openpyxl.chart.line_chart import LineChart
from urllib3.util import url
from tkinter import Button, Checkbutton, Label, Listbox, Scrollbar, Tk, messagebox, ttk
from tkinter.ttk import Combobox, Radiobutton
import multiprocessing
from _datetime import date, date
import os
import shutil
selected=[]
flag=-1
flag2=1
parameters=[]
processes=[]
lb=0

stockName='NIFTY'
titles={
    'A1':"CALLS",
    'I1':"PUTS",
    'A2':"OI",
    'B2':"COI",
    'C2':"Vol",
    'D2':"LTP",
    'E2':"CHNG",
    'G2':"Strike Price",
    'I2':"CHNG",
    'J2':"LTP",
    'K2':"Vol",
    'L2':"COI",
    'M2':"OI" ,
}
chartData={
    'Time':['COI','PRICE']
}
coiTitles={
    'A1':'CALLS',
    'K1':'PUTS',
    'A2':"OI",
    'B2':"VOL",
    'C2':"OI/VOL",
    'E2':"OI",
    'F2':"VOL",
    'G2':"OI/VOL",
    'I2':"STRIKE PRICE",
    'K2':"OI",
    'L2':"VOL",
    'M2':"OI/VOL",
    'O2':"OI",
    'P2':"VOL",
    'Q2':"OI/VOL",   
}
prevFile=''
firstFile=''
prevCopy=''
def getData(Time,counter):
    if(stockName=='NIFTY'or stockName=='BANKNIFTY'):
        url="https://www.nseindia.com/api/option-chain-indices?symbol="+stockName
    else:
        url="https://www.nseindia.com/api/option-chain-equities?symbol="+stockName
    headers={
        'User-Agent':'Chrome/87.0.4280.88' ,
    }
    while(1):
        try:
            res=requests.get(url,headers=headers)
            break
        except requests.exceptions.ConnectionError:
            root=Tk()
            root.withdraw()
            msgbox=messagebox.showerror('Error','Please check your connection re-trying in 5 seconds')
            root.destroy()  
            time.sleep(5)
    htmlContent=res.content
    soup=BeautifulSoup(htmlContent,'html.parser')
    while(1):
        try:
            data=json.loads(str(soup))
            break
        except json.decoder.JSONDecodeError:
            while(1):
                try:
                    res=requests.get(url,headers=headers)
                    break
                except requests.exceptions.ConnectionError:
                    root=Tk()
                    root.withdraw()
                    msgbox=messagebox.showerror('Error','Please check your connection re-trying in 5 seconds')
                    root.destroy()
                    time.sleep(5)
            htmlContent=res.content
            soup=BeautifulSoup(htmlContent,'html.parser')
    return(writeValues(data,Time,counter))
def createExcelFile(Path,sheetName):
      try:
            wb=Workbook(Path)
            ws=wb[sheetName]
            wb.close()
      except:
            wb=Workbook()
            ws=wb.create_sheet(sheetName)
            wb.save(Path)
            wb=openpyxl.load_workbook(Path)           
            wb.remove(wb['Sheet'])
            wb.save(Path)
            wb.close()
def writeValues(info,setTime,counter):
    Time=setTime[11:16]
    Time=Time.replace(':','.')
    price=info["records"]["underlyingValue"]
    Date=datetime.datetime.now()
    Date=Date.strftime('%d-%m-%Y')
    try:
        os.makedirs('C:\\'+stockName+"Options\\"+Date)
    except FileExistsError:
        pass
    path='C:\\'+stockName+"Options\\"+Date+"\\"+stockName+"-"+Time+" "+str(price)+" .xlsx"
    sheet=stockName
    titles['G1']=price
    createExcelFile(Path=path,sheetName=sheet)
    wb=openpyxl.load_workbook(path)
    ws=wb[sheet]
    sum1=0
    sum2=0
    select='CE'
    for title in titles:
        ws[title].value=titles[title]
    date=info["records"]["expiryDates"][0]
    ED=info["records"]['data'][0]['expiryDate']
    SP2=info["records"]['data'][0]['strikePrice']
    SP1=0
    count=1
    while(1):
        if(date==ED and SP1!=SP2):
            break
        SP1=SP2
        ED=info['records']['data'][count]['expiryDate']
        SP2=info['records']['data'][count]['strikePrice']
        count=count+1
    if(date==ED and SP2!=0):
        diff=int(SP2)-int(SP1)
    sub=diff*6 
    for i in range(2):
            row=3
            SP1=00
            for item in info["records"]["data"]:
                if select in item:
                        ED=item[select]["expiryDate"]                                                                                                          
                        SP=item[select]["strikePrice"]
                        OI=item[select]["openInterest"]
                        COI=item[select]["changeinOpenInterest"]
                        LTP=item[select]["lastPrice"]
                        VOL=item[select]["totalTradedVolume"]
                        UV=item[select]["underlyingValue"]
                        CHG=item[select]["change"]
                        if(int(UV)%diff==0):
                            min=int(UV-sub)
                            max=int(UV+sub)
                        else:
                            if(int(UV)%diff>=(diff/2)):
                                    UV=int(UV+(diff-(UV%diff)))
                                    min=UV-sub
                                    max=UV+sub
                            else:
                                    UV=int(UV-UV%diff)
                                    min=UV-sub
                                    max=UV+sub
                        if ED==date and SP>=min and SP<=max:
                            if select=='CE':
                                ws.cell(row=row,column=1).value=OI
                                ws.cell(row=row,column=2).value=COI
                                ws.cell(row=row,column=3).value=VOL
                                ws.cell(row=row,column=4).value=LTP
                                ws.cell(row=row,column=5).value=CHG
                                ws.cell(row=row,column=7).value=SP
                                sum1=sum1+COI
                            else:
                                ws.cell(row=row,column=9).value=CHG
                                ws.cell(row=row,column=10).value=LTP
                                ws.cell(row=row,column=11).value=VOL
                                ws.cell(row=row,column=12).value=COI
                                ws.cell(row=row,column=13).value=OI
                                sum2=sum2+COI
                            row+=1
            select='PE'
    ws['B18'].value=sum1
    ws['J18'].value=sum2
    ws["L18"].value=(sum1-sum2)
    wb.save(path)
    wb.close()
    wb1=openpyxl.load_workbook(path)
    wb1.close()
    chartData[setTime[11:16]]=[(sum1-sum2),price]
    global prevFile,firstFile,prevCopy
    if firstFile!=path and prevCopy!=path and firstFile!='':
        writeCoiChange(path) 
    prevFile=path
    if counter==1:
        firstFile=path
    return path
def writeCoiChange(Path):
    sheet='COI change'
    global prevFile,firstFile,prevCopy
    wb=openpyxl.load_workbook(Path)
    ws=wb[stockName]
    wb2=openpyxl.load_workbook(prevFile)
    ws2=wb2[stockName]
    wb3=openpyxl.load_workbook(firstFile)
    ws3=wb3[stockName]
    wb.create_sheet('COI change')
    wb.save(Path)
    ws4=wb['COI change']
    if(firstFile!=prevFile):
        for title in coiTitles:
                ws4[title].value=coiTitles[title]
    else:
        firstFileTitle={
                'A1':'CALLS',
                'G1':'PUTS',
                'A2':"OI",
                'B2':"VOL",
                'C2':"OI/VOL",
                'E2':"STRIKE PRICE",
                'G2':"OI",
                'H2':"VOL",
                'I2':"OI/VOL",   
            }
        for title in firstFileTitle:
                ws4[title].value=firstFileTitle[title]   
    if(firstFile!=prevFile):
        ws4['G1'].value=(prevFile[6:11].replace('.',':'))
        ws4['Q1'].value=(prevFile[6:11].replace('.',':'))
        ws4['M1'].value=(firstFile[6:11].replace('.',':'))
    else:
        ws4['I1'].value=(firstFile[6:11].replace('.',':'))
    ws4['C1'].value=(firstFile[6:11].replace('.',':'))
    fi=ws.max_row
    OIcol=1
    VOLcol=3
    col=1
    tempWs=ws3
    for count in range (4):
        row=3
        for i in range(3,fi-2):
            SP1=tempWs.cell(row=i,column=7).value
            OI1=int(tempWs.cell(row=i,column=OIcol).value)
            VOL1=int(tempWs.cell(row=i,column=VOLcol).value)
            for j in range(3,fi-2):
                    OI2=(int)(ws.cell(row=j,column=OIcol).value)
                    VOL2=int(ws.cell(row=j,column=VOLcol).value)
                    SP2=ws.cell(row=j,column=7).value
                    if(SP1==SP2):#pre
                        ws4.cell(row,col+1).value=VOL2-VOL1
                        ws4.cell(row,col).value=OI2-OI1
                        if(VOL2-VOL1!=0):
                                ws4.cell(row,col+2).value=((OI2-OI1)/(VOL2-VOL1))
                        else:
                                ws4.cell(row,col+2).value='∞'
                        if(count==0 and firstFile!=prevFile):
                            ws4.cell(row,9).value=SP2
                        elif(count==0 and firstFile==prevFile):
                            ws4.cell(row,5).value=SP2
                        row=row+1
        if(prevFile==firstFile):
            count=count+1
            col=7
        else:
            col=((col+4,col+6))[col+4==9]
            tempWs=(ws3,ws2)[count%2==0]
        if(count==1):
            OIcol=13
            VOLcol=11   
    wb.save(Path)
    wb.close()
    wb2.close()
    wb3.close()
    prevCopy=prevFile
    createChart()
def createChart():
    Date=datetime.datetime.now()
    Date=Date.strftime('%d-%m-%Y')
    path='C:\\'+stockName+"Options\\"+Date+"\\"+stockName+'Chart.xlsx'
    sheet=stockName+'Chart'
    createExcelFile(Path=path,sheetName=sheet)
    wb=openpyxl.load_workbook(path)
    wb.remove(wb[sheet])
    wb.create_sheet(sheet)
    wb.save(path)
    ws=wb[sheet]
    row=1
    for key,values in chartData.items():
        ws.cell(row=row,column=1).value=key
        ws.cell(row=row,column=2).value=values[0]
        ws.cell(row=row,column=3).value=values[1]
        row=row+1
    wb.save(path)
    categories=Reference(ws,min_col=1,min_row=2,max_col=1,max_row=row)
    values=Reference(ws,min_col=2,min_row=2,max_col=2,max_row=row)
    chart=LineChart()
    ws.add_chart(chart,"E2")
    chart.title=stockName
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.title="Time"
    chart.y_axis.title="COI"
    chart.add_data(values)
    chart.add_data(categories)
    chart.set_categories(categories)
    wb.save(path)
    wb.close()

def IsSelected(txt):
    global flag,flag2,stockName,parameters,selected
    if txt not in selected and flag2==1:
        selected.append(txt)
        parameters=selected
        parameters.sort()
        flag=1
        flag2=0
        stockName=txt
        return txt
    elif flag2==0:
        msgbox=messagebox.showerror('Error',"Please select Refresh period of previous stock before selecting "+txt)
    else:
        msgbox=messagebox.showerror('Error',txt+" is already selected")
    return
def refereshPeriod(t):
    global flag,flag2,processes
    if(flag==1):
        flag=-1
        flag2=1
        period=t*60
        p=multiprocessing.Process(target=autoRefresh,args=[period,stockName])
        processes.append(p)
        p.start()
    else:
        msgbox=messagebox.showerror('Error',"Please select a stock")
def Delete():
    global lb,parameters,selected,processes
    flag=0
    if(len(lb.curselection())>0 and (len(parameters))>0):
        for item in reversed(lb.curselection()):
            lb.delete(item)
            temp=parameters[item]
            parameters.remove(parameters[item])
            for i in range(0,len(selected)):
                if temp == selected[i]:
                    selected.remove(temp)
                    p=processes[i]
                    shutil.rmtree('C:\\'+stockName+"Options")
                    p.terminate()
                    processes.remove(p)
            flag=1
    else:
        msgbox=messagebox.showerror('Error','please select the stock you want to remove.')
    if(flag==1):
        flag=Refresh(flag)
def Refresh(flag):
    global parameters,lb
    if(flag==1):
        lb.delete(0,'end')
        for i in parameters:
            lb.insert('end',i)
        return 0
def listbox():
    global parameters,lb
    wd=Tk()
    wd.title('selected stock list')
    wd.geometry('350x300')
    lb=Listbox(wd,bg='white',selectmode='multiple')
    scrollbar=Scrollbar(wd)
    for i in parameters:
        lb.insert('end',i)
    lb.place(x=30,y=10,height=220,width=250)
    scrollbar.place(x=275,y=10,height=220,width=15)
    lb.config(yscrollcommand = scrollbar.set) 
    scrollbar.config(command=lb.yview)
    wd.resizable(0,0)
    btn=Button(wd,text='STOP!',bg='#e2d1d0',command=Delete)
    btn.place(x=140,y=250,height=35,width=50)
    wd.mainloop()  
def Choice():
    window=Tk()
    window.geometry('350x400')
    l1=Label(window,text='View Option Contract for:',bg='#f2f1cc')
    l1.place(x=30,y=35,height=35,width=150)
    b1=Button(window,text="NIFTY",font=('Arial Bold',10),)
    b1.configure(command=lambda :IsSelected(b1['text']),bg='#e2d1d0')
    b1.place(x=180,y=35,height=35,width=100)
    b2=Button(window,text="BANKNIFTY",font=('Arial Bold',10))
    l2=Label(window,text='View Option Contract for:',bg='#f2f1cc')
    l2.place(x=30,y=110,height=35,width=150)
    b2.configure(command=lambda:IsSelected(b2['text']),bg='#e2d1d0')
    b2.place(x=180,y=110,height=35,width=100)
    l3=Label(window,text='Select Symbol',font=('Arial',10),bg='#f2f1cc')
    l3.place(x=60,y=185,height=35,width=100)
    cb=Combobox(window)
    cb['values']=["AARTIIND","ACC","ADANIENT","ADANIPORTS","AMARAJABAT","AMBUJACEM","APOLLOHOSP","APOLLOTYRE","ASHOKLEY","ASIANPAINT","AUROPHARMA","AXISBANK","BAJAJ-AUTO","BAJAJFINSV","BAJFINANCE",
    "BALKRISIND","BANDHANBNK","BANKBARODA","BATAINDIA","BEL","BERGEPAINT","BHARATFORG","BHARTIARTL","BHEL","BIOCON","BOSCHLTD","BPCL","BRITANNIA",
    "CADILAHC","CANBK","CHOLAFIN","CIPLA","COALINDIA","COFORGE","COLPAL","CONCOR","CUMMINSIND","DABUR","DIVISLAB","DLF","DRREDDY","EICHERMOT",
    "ESCORTS","EXIDEIND","FEDERALBNK","GAIL","GLENMARK","GMRINFRA","GODREJCP","GODREJPROP","GRASIM","HAVELLS","HCLTECH","HDFC","HDFCAMC","HDFCBANK","HDFCLIFE",
    "HEROMOTOCO","HINDALCO","HINDPETRO","HINDUNILVR","IBULHSGFIN","ICICIBANK","ICICIGI","ICICIPRULI","IDEA","IDFCFIRSTB","IGL","INDIGO","INDUSINDBK",
    "INDUSTOWER","INFRATEL","INFY","IOC","ITC","JINDALSTEL","JSWSTEEL","JUBLFOOD","KOTAKBANK","L&TFH","LALPATHLAB","LICHSGFIN","LT","LUPIN","M&M",
    "M&MFIN","MANAPPURAM","MARICO","MARUTI","MCDOWELL-N","MFSL","MGL","MINDTREE","MOTHERSUMI","MRF","MUTHOOTFIN","NATIONALUM","NAUKRI","NESTLEIND",
    "NMDC","NTPC","ONGC","PAGEIND","PEL","PETRONET","PFC","PIDILITIND","PNB","POWERGRID","PVR","RAMCOCEM","RBLBANK","RECLTD","RELIANCE","SAIL",
    "SBILIFE","SBIN","SHREECEM","SIEMENS","SRF","SRTRANSFIN","SUNPHARMA","SUNTV","TATACHEM","TATACONSUM","TATAMOTORS","TATAPOWER","TATASTEEL","TCS",
    "TECHM","TITAN","TORNTPHARM","TORNTPOWER","TVSMOTOR","UBL","ULTRACEMCO","UPL","VEDL","ZEEL","VOLTAS",]
    cb.current(0)
    cb.place(x=160,y=185,height=35,width=100)
    l4=Label(window,text="Refresh in (minutes):",font=('Arial',10),bg="#f2f1cc")
    l4.place(x=25,y=260,height=40,width=130)
    rad1=Radiobutton(window,text='5',value=1,command=lambda:refereshPeriod(5))
    rad2=Radiobutton(window,text='15',value=0,command=lambda:refereshPeriod(15))
    rad3=Radiobutton(window,text='30',value=3,command=lambda:refereshPeriod(30))
    rad4=Radiobutton(window,text='60',value=4,command=lambda:refereshPeriod(60))
    rad1.place(x=160,y=260,height=40,width=30)
    rad2.place(x=200,y=260,height=40,width=35)
    rad3.place(x=240,y=260,height=40,width=35)
    rad4.place(x=280,y=260,height=40,width=35)
    l5=Label(window,text="For non-commerical use only.",font=('Arial',10),fg='red')
    l5.place(x=75,y=375,height=35,width=175)
    cb.bind("<<ComboboxSelected>>", lambda Combobox=cb:IsSelected(cb.get()))
    b3=Button(window,text="View stocks selected.",font=('Arial',10),bg='#e2d1d0',command=listbox)
    b3.place(x=100,y=325,height=35,width=150)
    window.title('Stock data updater')
    window.resizable(0,0)
    window.mainloop()
def autoRefresh(interval,stName):
    global stockName
    path=''
    stockName=stName
    Date=datetime.datetime.now()
    Date=Date.strftime('%d-%m-%Y')
    chartPath='C:\\'+stockName+"Options\\"+Date+"\\"+stockName+'Chart.xlsx'
    newPath=stockName+"Options\\"+Date
    count=1
    getTime=str(datetime.datetime.now())
    min=int(getTime[14:16])
    hr=int(getTime[11:13])
    sec=int(getTime[17:19])
    try:
        os.makedirs(newPath)
    except FileExistsError:
        pass
    if getTime[11:16]=="09:20":
            count=1
            path=getData(getTime,count)
            shutil.copy(path,newPath)
            try:
                shutil.copy(chartPath,newPath)
            except FileNotFoundError:
                pass
    
    while (int(min)%(interval/60)!=0):
        time.sleep(60-sec)
        getTime=str(datetime.datetime.now())
        min=int(getTime[14:16]) 
        sec=int(getTime[17:19])
    while hr<=15:
        getTime=str(datetime.datetime.now())
        sec=int(getTime[17:19])
        path=getData(getTime,count)
        shutil.copy(path,newPath)
        try:
            shutil.copy(chartPath,newPath)
        except FileNotFoundError:
            pass
        time.sleep(interval)
        count=count+1  
        hr=int(getTime[11:13])
    if(hr>15):
        count=1
        path=getData(getTime,count)
        shutil.copy(path,newPath)
        try:
            shutil.copy(chartPath,newPath)
        except FileNotFoundError:
            pass
    shutil.rmtree('C:\\'+stockName+"Options")
if __name__ == "__main__":
    Choice()
